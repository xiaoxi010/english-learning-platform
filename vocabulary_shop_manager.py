# vocabulary_manager.py - 完整版本（添加Excel导入导出）
import json
import os
import sqlite3
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook, load_workbook

# 新词汇库固定放在项目目录，避免 Flask 工作目录变化导致读到别的 db 文件
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_DB_PATH = os.path.join(_BASE_DIR, 'vocabulary_shop.db')
_LEGACY_DB_PATH = os.path.join(_BASE_DIR, 'vocabulary_new.db')
LEGACY_SYSTEM_DICTS = frozenset({'默认词典', 'A-错题词典'})


class VocabularyManager:
    def __init__(self, db_path=None):
        if db_path is None and not os.path.exists(DEFAULT_DB_PATH) and os.path.exists(_LEGACY_DB_PATH):
            os.rename(_LEGACY_DB_PATH, DEFAULT_DB_PATH)
        self.db_path = os.path.abspath(db_path or DEFAULT_DB_PATH)
        self.current_dictionary = None
        is_new_db = not os.path.exists(self.db_path)
        self.init_database()
        if is_new_db:
            print(f'已创建空词汇数据库: {self.db_path}')
        else:
            self._remove_legacy_system_dicts()

    def init_database(self):
        """初始化数据库表结构（不预置任何词典）"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # 创建词典表
                # 创建词典表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS dictionaries
            (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dict_name TEXT UNIQUE NOT NULL,
                created_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active BOOLEAN DEFAULT 1,
                description TEXT,
                sort_order INTEGER DEFAULT 0
            )
        ''')

        # 创建单词组表
        cursor.execute('''
                       CREATE TABLE IF NOT EXISTS word_groups
                       (
                           id
                           INTEGER
                           PRIMARY
                           KEY
                           AUTOINCREMENT,
                           group_name
                           TEXT
                           NOT
                           NULL,
                           dict_id
                           INTEGER
                           NOT
                           NULL,
                           created_time
                           TIMESTAMP
                           DEFAULT
                           CURRENT_TIMESTAMP,
                           description
                           TEXT,
                           FOREIGN
                           KEY
                       (
                           dict_id
                       ) REFERENCES dictionaries
                       (
                           id
                       ),
                           UNIQUE
                       (
                           group_name,
                           dict_id
                       )
                           )
                       ''')

        # 创建单词表
        cursor.execute('''
                       CREATE TABLE IF NOT EXISTS words
                       (
                           id
                           INTEGER
                           PRIMARY
                           KEY
                           AUTOINCREMENT,
                           word
                           TEXT
                           NOT
                           NULL,
                           part_of_speech
                           TEXT
                           NOT
                           NULL,
                           chinese_meaning
                           TEXT
                           NOT
                           NULL,
                           group_id
                           INTEGER,
                           created_time
                           TIMESTAMP
                           DEFAULT
                           CURRENT_TIMESTAMP,
                           review_count
                           INTEGER
                           DEFAULT
                           0,
                           last_reviewed
                           TIMESTAMP,
                           FOREIGN
                           KEY
                       (
                           group_id
                       ) REFERENCES word_groups
                       (
                           id
                       ),
                           UNIQUE
                       (
                           word,
                           group_id
                       )
                           )
                       ''')

        # 创建索引
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_word ON words(word)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_group_id ON words(group_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_group_name ON word_groups(group_name)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_dict_id ON word_groups(dict_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_dict_name ON dictionaries(dict_name)')

        conn.commit()
        conn.close()
        self._add_sort_order_column()
        self._ensure_shop_columns()

    def _ensure_shop_columns(self):
        """商城字段：积分价格、图标（默认积分 0）"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('PRAGMA table_info(dictionaries)')
            columns = {column[1] for column in cursor.fetchall()}
            if 'price' not in columns:
                cursor.execute('ALTER TABLE dictionaries ADD COLUMN price INTEGER DEFAULT 0')
            if 'icon' not in columns:
                cursor.execute('ALTER TABLE dictionaries ADD COLUMN icon TEXT DEFAULT "📚"')
            cursor.execute('UPDATE dictionaries SET price = 0 WHERE price IS NULL')
            cursor.execute('UPDATE dictionaries SET icon = "📚" WHERE icon IS NULL OR icon = ""')
            conn.commit()
            conn.close()
        except Exception as e:
            print(f'添加商城字段失败: {e}')

    def get_shop_catalog(self) -> List[Dict]:
        """获取商城上架词典（含词数、积分）"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.id, d.dict_name, d.description,
                       COALESCE(d.price, 0), COALESCE(d.icon, '📚'),
                       COUNT(w.id) AS word_count
                FROM dictionaries d
                LEFT JOIN word_groups wg ON wg.dict_id = d.id
                LEFT JOIN words w ON w.group_id = wg.id
                WHERE d.is_active = 1
                GROUP BY d.id
                ORDER BY d.sort_order, d.created_time DESC
            ''')
            items = []
            for row in cursor.fetchall():
                items.append({
                    'id': row[0],
                    'dict_name': row[1],
                    'description': row[2] or '',
                    'price': row[3] or 0,
                    'icon': row[4] or '📚',
                    'word_count': row[5] or 0,
                })
            conn.close()
            return items
        except Exception as e:
            print(f'获取商城目录失败: {e}')
            return []

    def get_dictionary_for_shop(self, dict_name: str) -> Optional[Dict]:
        """获取商城中指定词典信息"""
        for item in self.get_shop_catalog():
            if item['dict_name'] == dict_name:
                return item
        return None

    def get_dictionary_groups_with_words(self, dict_name: str) -> List[Dict]:
        """获取词典下所有词组及单词（用于导入）"""
        groups = []
        for group in self.get_all_groups_sorted('name'):
            if group['dict_name'] != dict_name:
                continue
            detail = self.get_word_group(group['group_name'], dict_name)
            if detail:
                groups.append(detail)
        return groups

    def _remove_legacy_system_dicts(self):
        """清理旧版系统词典（默认词典、错题词典）"""
        for name in LEGACY_SYSTEM_DICTS:
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM dictionaries WHERE dict_name = ?', (name,))
                row = cursor.fetchone()
                if not row:
                    conn.close()
                    continue
                dict_id = row[0]
                cursor.execute(
                    'DELETE FROM words WHERE group_id IN (SELECT id FROM word_groups WHERE dict_id = ?)',
                    (dict_id,)
                )
                cursor.execute('DELETE FROM word_groups WHERE dict_id = ?', (dict_id,))
                cursor.execute('DELETE FROM dictionaries WHERE id = ?', (dict_id,))
                conn.commit()
                conn.close()
            except Exception as e:
                print(f"清理系统词典 {name} 失败: {e}")

    def init_wrong_book_system(self):
        """初始化错题本系统"""
        try:
            # 确保错题词典始终为激活状态且不可手动修改
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # 确保错题词典存在且激活
            cursor.execute('''
                INSERT OR IGNORE INTO dictionaries (dict_name, description, is_active, sort_order) 
                VALUES (?, ?, 1, -1)
            ''', ('A-错题词典', '系统错题词典，自动管理'))

            # 强制设置错题词典为激活状态
            cursor.execute('''
                UPDATE dictionaries SET is_active = 1, sort_order = -1 
                WHERE dict_name = 'A-错题词典'
            ''')

            conn.commit()
            conn.close()

            # 创建10个错题本（如果不存在）
            for i in range(1, 11):
                group_name = f'A-错题本-{i}'
                self._create_wrong_book_if_not_exists(group_name)

            print("错题本系统初始化完成")

        except Exception as e:
            print(f"初始化错题本系统失败: {e}")

    def _create_wrong_book_if_not_exists(self, group_name: str) -> bool:
        """创建错题本（如果不存在）"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # 检查错题本是否已存在
            cursor.execute('''
                SELECT COUNT(*) FROM word_groups wg
                JOIN dictionaries d ON wg.dict_id = d.id
                WHERE wg.group_name = ? AND d.dict_name = 'A-错题词典'
            ''', (group_name,))

            if cursor.fetchone()[0] == 0:
                # 获取错题词典的ID
                cursor.execute("SELECT id FROM dictionaries WHERE dict_name = 'A-错题词典'")
                dict_result = cursor.fetchone()
                if dict_result:
                    dict_id = dict_result[0]
                    cursor.execute(
                        'INSERT INTO word_groups (group_name, dict_id, description) VALUES (?, ?, ?)',
                        (group_name, dict_id, '系统错题本，自动管理')
                    )
                    conn.commit()
                    print(f"已创建 {group_name}")

            conn.close()
            return True

        except Exception as e:
            print(f"创建错题本失败: {e}")
            return False

    def _add_sort_order_column(self):
        """检查并添加 sort_order 列"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # 检查是否存在 sort_order 列
            cursor.execute("PRAGMA table_info(dictionaries)")
            columns = [column[1] for column in cursor.fetchall()]

            if 'sort_order' not in columns:
                print("添加 sort_order 列到 dictionaries 表...")
                cursor.execute('ALTER TABLE dictionaries ADD COLUMN sort_order INTEGER DEFAULT 0')

                # 为现有数据设置排序值
                cursor.execute('SELECT id FROM dictionaries ORDER BY created_time')
                rows = cursor.fetchall()
                for order, row in enumerate(rows):
                    cursor.execute('UPDATE dictionaries SET sort_order = ? WHERE id = ?', (order, row[0]))

                conn.commit()
                print("sort_order 列添加成功")

            conn.close()
        except Exception as e:
            print(f"添加 sort_order 列失败: {e}")

    def update_dictionary_order(self, dict_names: List[str]) -> bool:
        """更新词典顺序"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # 为每个词典更新排序字段
            for order, dict_name in enumerate(dict_names):
                cursor.execute(
                    'UPDATE dictionaries SET sort_order = ? WHERE dict_name = ?',
                    (order, dict_name)
                )

            conn.commit()
            conn.close()
            return True

        except Exception as e:
            print(f"更新词典顺序失败: {e}")
            return False

    def get_dictionaries_ordered(self) -> List[Dict]:
        """获取按顺序排列的词典列表"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                           SELECT id,
                                  dict_name,
                                  created_time,
                                  is_active,
                                  description,
                                  sort_order,
                                  (SELECT COUNT(*) FROM word_groups WHERE dict_id = dictionaries.id) as group_count
                           FROM dictionaries
                           ORDER BY sort_order, created_time DESC
                           ''')

            dictionaries = []
            for row in cursor.fetchall():
                dictionaries.append({
                    'id': row[0],
                    'dict_name': row[1],
                    'created_time': row[2],
                    'is_active': bool(row[3]),
                    'description': row[4],
                    'sort_order': row[5],
                    'group_count': row[6]
                })

            conn.close()
            return dictionaries

        except Exception as e:
            print(f"获取有序词典列表失败: {e}")
            return []

    def get_all_groups_sorted(self, sort_by: str = "name") -> List[Dict]:
        """获取排序后的单词组列表"""
        try:
            active_dicts = self.get_active_dictionaries()
            if not active_dicts:
                return []

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            placeholders = ','.join(['?'] * len(active_dicts))

            if sort_by == "name":
                # 按组名排序（字母数字顺序）
                cursor.execute(f'''
                    SELECT wg.group_name, wg.created_time, wg.description, 
                           COUNT(w.id) as word_count, d.dict_name
                    FROM word_groups wg
                    JOIN dictionaries d ON wg.dict_id = d.id
                    LEFT JOIN words w ON wg.id = w.group_id
                    WHERE d.dict_name IN ({placeholders})
                    GROUP BY wg.id
                    ORDER BY wg.group_name COLLATE NOCASE
                ''', active_dicts)
            elif sort_by == "date":
                # 按创建时间排序 - 使用ID倒序（通常ID越大越新）
                cursor.execute(f'''
                        SELECT wg.group_name, wg.created_time, wg.description, 
                               COUNT(w.id) as word_count, d.dict_name
                        FROM word_groups wg
                        JOIN dictionaries d ON wg.dict_id = d.id
                        LEFT JOIN words w ON wg.id = w.group_id
                        WHERE d.dict_name IN ({placeholders})
                        GROUP BY wg.id
                        ORDER BY wg.id DESC  -- 使用ID倒序
                ''', active_dicts)
            elif sort_by == "dict":
                # 按词典排序
                cursor.execute(f'''
                    SELECT wg.group_name, wg.created_time, wg.description, 
                           COUNT(w.id) as word_count, d.dict_name
                    FROM word_groups wg
                    JOIN dictionaries d ON wg.dict_id = d.id
                    LEFT JOIN words w ON wg.id = w.group_id
                    WHERE d.dict_name IN ({placeholders})
                    GROUP BY wg.id
                    ORDER BY d.dict_name, wg.group_name
                ''', active_dicts)
            else:
                # 默认按组名排序
                cursor.execute(f'''
                    SELECT wg.group_name, wg.created_time, wg.description, 
                           COUNT(w.id) as word_count, d.dict_name
                    FROM word_groups wg
                    JOIN dictionaries d ON wg.dict_id = d.id
                    LEFT JOIN words w ON wg.id = w.group_id
                    WHERE d.dict_name IN ({placeholders})
                    GROUP BY wg.id
                    ORDER BY wg.group_name COLLATE NOCASE
                ''', active_dicts)

            groups = []
            for row in cursor.fetchall():
                groups.append({
                    'group_name': row[0],
                    'created_time': row[1],
                    'description': row[2],
                    'word_count': row[3],
                    'dict_name': row[4]
                })

            conn.close()
            return groups

        except Exception as e:
            print(f"获取排序单词组列表失败: {e}")
            return []

    # 以下保持原有方法，但修改 get_all_groups 调用 get_all_groups_sorted
    def get_all_groups(self, sort_by: str = "name") -> List[Dict]:
        """获取所有激活词典的单词组列表（保持兼容性）"""
        return self.get_all_groups_sorted(sort_by)

    def add_dictionary(self, dict_name: str, description: str = "") -> bool:
        """添加词典"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # 获取当前最大排序值
            cursor.execute('SELECT MAX(sort_order) FROM dictionaries')
            max_order = cursor.fetchone()[0] or 0

            cursor.execute(
                'INSERT INTO dictionaries (dict_name, description, sort_order) VALUES (?, ?, ?)',
                (dict_name, description, max_order + 1)
            )

            conn.commit()
            conn.close()
            return True

        except Exception as e:
            print(f"添加词典失败: {e}")
            return False

    def dictionary_exists(self, dict_name: str) -> bool:
        """检查词典是否存在"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT 1 FROM dictionaries WHERE dict_name = ?', (dict_name,))
            exists = cursor.fetchone() is not None
            conn.close()
            return exists
        except Exception as e:
            print(f"检查词典是否存在失败: {e}")
            return False

    def ensure_dictionary(self, dict_name: str) -> bool:
        """若词典不存在则自动创建"""
        if not dict_name:
            return False
        if self.dictionary_exists(dict_name):
            return True
        return self.add_dictionary(dict_name)

    def get_dictionary_status(self) -> Dict[str, bool]:
        """获取所有词典的使用状态"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('SELECT dict_name, is_active FROM dictionaries ORDER BY sort_order')
            status = {}
            for row in cursor.fetchall():
                status[row[0]] = bool(row[1])

            conn.close()
            return status
        except Exception as e:
            print(f"获取词典状态失败: {e}")
            return {}

    def set_dictionary_status(self, dict_name: str, is_active: bool) -> bool:
        """设置词典使用状态"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute(
                'UPDATE dictionaries SET is_active = ? WHERE dict_name = ?',
                (1 if is_active else 0, dict_name)
            )

            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"设置词典状态失败: {e}")
            return False

    def get_active_dictionaries(self) -> List[str]:
        """获取所有被激活的词典名称"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                SELECT dict_name FROM dictionaries 
                WHERE is_active = 1
                ORDER BY sort_order
            ''')
            active_dicts = [row[0] for row in cursor.fetchall()]

            conn.close()
            return active_dicts
        except Exception as e:
            print(f"获取激活词典失败: {e}")
            return []

    def get_current_wrong_book(self) -> Optional[str]:
        """获取当前可添加单词的错题本（第一个未满50个单词的）"""
        try:
            for i in range(1, 11):
                group_name = f'A-错题本-{i}'
                group_data = self.get_word_group(group_name, 'A-错题词典')

                if group_data:
                    word_count = len(group_data.get('words', []))
                    if word_count < 50:
                        return group_name

            # 所有错题本都满了
            return None

        except Exception as e:
            print(f"获取当前错题本失败: {e}")
            return None

    # 添加新方法：获取错题本统计信息
    def get_wrong_book_statistics(self) -> Dict:
        """获取错题本统计信息"""
        try:
            stats = {
                'total_wrong_words': 0,
                'books': []
            }

            for i in range(1, 11):
                group_name = f'A-错题本-{i}'
                group_data = self.get_word_group(group_name, 'A-错题词典')

                if group_data:
                    word_count = len(group_data.get('words', []))
                    stats['total_wrong_words'] += word_count
                    stats['books'].append({
                        'name': group_name,
                        'word_count': word_count,
                        'is_full': word_count >= 50
                    })

            return stats

        except Exception as e:
            print(f"获取错题本统计失败: {e}")
            return {'total_wrong_words': 0, 'books': []}

    # 添加核心方法：添加单词到错题本
    def add_word_to_wrong_book(self, word: str, part_of_speech: str, chinese_meaning: str) -> bool:
        """将单词添加到错题本（自动选择未满的错题本）"""
        try:
            # 获取当前可用的错题本
            current_book = self.get_current_wrong_book()
            if not current_book:
                print("所有错题本已满，无法添加")
                return False

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # 获取错题本的group_id
            cursor.execute('''
                SELECT wg.id
                FROM word_groups wg
                JOIN dictionaries d ON wg.dict_id = d.id
                WHERE wg.group_name = ? AND d.dict_name = 'A-错题词典'
            ''', (current_book,))

            result = cursor.fetchone()
            if not result:
                conn.close()
                return False

            group_id = result[0]

            # 检查该单词是否已在任何错题本中存在
            cursor.execute('''
                SELECT w.id FROM words w
                JOIN word_groups wg ON w.group_id = wg.id
                JOIN dictionaries d ON wg.dict_id = d.id
                WHERE w.word = ? AND d.dict_name = 'A-错题词典'
            ''', (word,))

            if cursor.fetchone():
                # 单词已存在于某个错题本中，不重复添加
                print(f"单词 '{word}' 已在错题本中存在，跳过添加")
                conn.close()
                return True

            # 添加单词
            cursor.execute('''
                INSERT INTO words (word, part_of_speech, chinese_meaning, group_id)
                VALUES (?, ?, ?, ?)
            ''', (word, part_of_speech, chinese_meaning, group_id))

            conn.commit()
            conn.close()

            print(f"单词 '{word}' 已添加到 {current_book}")
            return True

        except Exception as e:
            print(f"添加单词到错题本失败: {e}")
            return False

    # 添加核心方法：从错题本删除单词
    def remove_word_from_wrong_book(self, word: str) -> bool:
        """从错题本中删除单词（从所有错题本中查找并删除）"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                DELETE FROM words 
                WHERE word = ? AND group_id IN (
                    SELECT wg.id 
                    FROM word_groups wg
                    JOIN dictionaries d ON wg.dict_id = d.id
                    WHERE d.dict_name = 'A-错题词典'
                )
            ''', (word,))

            rows_deleted = cursor.rowcount
            conn.commit()
            conn.close()

            if rows_deleted > 0:
                print(f"单词 '{word}' 已从错题本中删除")
            return rows_deleted > 0

        except Exception as e:
            print(f"从错题本删除单词失败: {e}")
            return False

    # 添加新方法：清理空的错题本（整理碎片）
    def compact_wrong_books(self) -> bool:
        """整理错题本，将单词向前移动，删除空本"""
        try:
            all_wrong_words = []

            # 收集所有错题本的单词
            for i in range(1, 11):
                group_name = f'A-错题本-{i}'
                group_data = self.get_word_group(group_name, 'A-错题词典')
                if group_data and group_data['words']:
                    for word in group_data['words']:
                        all_wrong_words.append({
                            'word': word['word'],
                            'part_of_speech': word['part_of_speech'],
                            'chinese_meaning': word['chinese_meaning']
                        })

            # 清空所有错题本
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                DELETE FROM words WHERE group_id IN (
                    SELECT wg.id FROM word_groups wg
                    JOIN dictionaries d ON wg.dict_id = d.id
                    WHERE d.dict_name = 'A-错题词典'
                )
            ''')
            conn.commit()
            conn.close()

            # 按顺序重新添加单词
            for i, word_data in enumerate(all_wrong_words):
                book_index = (i // 50) + 1
                if book_index > 10:
                    print(f"警告：单词数量超过500个，部分单词将丢失")
                    break

                self.add_word_to_wrong_book(
                    word_data['word'],
                    word_data['part_of_speech'],
                    word_data['chinese_meaning']
                )

            print(f"错题本整理完成，共 {len(all_wrong_words[:500])} 个单词")
            return True

        except Exception as e:
            print(f"整理错题本失败: {e}")
            return False

    def delete_dictionary(self, dict_name: str) -> bool:
        """删除词典"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM dictionaries WHERE dict_name = ?', (dict_name,))
            dict_id_result = cursor.fetchone()
            if not dict_id_result:
                return False

            dict_id = dict_id_result[0]

            # 删除词典下的所有单词
            cursor.execute('DELETE FROM words WHERE group_id IN (SELECT id FROM word_groups WHERE dict_id = ?)',
                           (dict_id,))

            # 删除词典下的所有单词组
            cursor.execute('DELETE FROM word_groups WHERE dict_id = ?', (dict_id,))

            # 删除词典
            cursor.execute('DELETE FROM dictionaries WHERE id = ?', (dict_id,))

            conn.commit()
            conn.close()

            if self.current_dictionary == dict_name:
                self.current_dictionary = None

            return True

        except Exception as e:
            print(f"删除词典失败: {e}")
            return False

    def add_word_group(self, group_name: str, words_data: List[Dict], description: str = "",
                       dict_name: str = None) -> bool:
        """添加单词组"""
        try:
            if not dict_name:
                return False

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM dictionaries WHERE dict_name = ?', (dict_name,))
            dict_result = cursor.fetchone()
            if not dict_result:
                conn.close()
                return False
            dict_id = dict_result[0]

            cursor.execute(
                'INSERT OR REPLACE INTO word_groups (group_name, dict_id, description) VALUES (?, ?, ?)',
                (group_name, dict_id, description)
            )

            cursor.execute('SELECT id FROM word_groups WHERE group_name = ? AND dict_id = ?', (group_name, dict_id))
            group_id = cursor.fetchone()[0]

            for word_data in words_data:
                cursor.execute('''
                    INSERT OR REPLACE INTO words 
                    (word, part_of_speech, chinese_meaning, group_id) 
                    VALUES (?, ?, ?, ?)
                ''', (
                    word_data['word'],
                    word_data['part_of_speech'],
                    word_data['chinese_meaning'],
                    group_id
                ))

            conn.commit()
            conn.close()
            return True

        except Exception as e:
            print(f"添加单词组失败: {e}")
            return False

    def get_word_group(self, group_name: str, dict_name: str = None) -> Optional[Dict]:
        """获取指定单词组的详细信息"""
        try:
            if not dict_name:
                return None

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                           SELECT wg.id, wg.group_name, wg.created_time, wg.description, d.dict_name
                           FROM word_groups wg
                                    JOIN dictionaries d ON wg.dict_id = d.id
                           WHERE wg.group_name = ?
                             AND d.dict_name = ?
                           ''', (group_name, dict_name))

            group_info = cursor.fetchone()

            if not group_info:
                return None

            cursor.execute('''
                           SELECT word, part_of_speech, chinese_meaning, review_count, last_reviewed
                           FROM words
                           WHERE group_id = ?
                           ORDER BY word COLLATE NOCASE
                           ''', (group_info[0],))

            words = []
            for row in cursor.fetchall():
                words.append({
                    'word': row[0],
                    'part_of_speech': row[1],
                    'chinese_meaning': row[2],
                    'review_count': row[3],
                    'last_reviewed': row[4]
                })

            conn.close()

            return {
                'group_id': group_info[0],
                'group_name': group_info[1],
                'created_time': group_info[2],
                'description': group_info[3],
                'dict_name': group_info[4],
                'words': words,
                'word_count': len(words)
            }

        except Exception as e:
            print(f"获取单词组失败: {e}")
            return None

    def search_words(self, keyword: str, search_by: str = "word") -> List[Dict]:
        """搜索单词（只在激活的词典中搜索）"""
        try:
            active_dicts = self.get_active_dictionaries()
            if not active_dicts:
                return []

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            placeholders = ','.join(['?'] * len(active_dicts))

            if search_by == "word":
                cursor.execute(f'''
                    SELECT w.word, w.part_of_speech, w.chinese_meaning, wg.group_name, d.dict_name
                    FROM words w
                    JOIN word_groups wg ON w.group_id = wg.id
                    JOIN dictionaries d ON wg.dict_id = d.id
                    WHERE w.word LIKE ? AND d.dict_name IN ({placeholders})
                    ORDER BY w.word COLLATE NOCASE
                ''', (f'%{keyword}%', *active_dicts))
            elif search_by == "meaning":
                cursor.execute(f'''
                    SELECT w.word, w.part_of_speech, w.chinese_meaning, wg.group_name, d.dict_name
                    FROM words w
                    JOIN word_groups wg ON w.group_id = wg.id
                    JOIN dictionaries d ON wg.dict_id = d.id
                    WHERE w.chinese_meaning LIKE ? AND d.dict_name IN ({placeholders})
                    ORDER BY w.word COLLATE NOCASE
                ''', (f'%{keyword}%', *active_dicts))
            else:
                cursor.execute(f'''
                    SELECT w.word, w.part_of_speech, w.chinese_meaning, wg.group_name, d.dict_name
                    FROM words w
                    JOIN word_groups wg ON w.group_id = wg.id
                    JOIN dictionaries d ON wg.dict_id = d.id
                    WHERE (w.word LIKE ? OR w.chinese_meaning LIKE ?) AND d.dict_name IN ({placeholders})
                    ORDER BY w.word COLLATE NOCASE
                ''', (f'%{keyword}%', f'%{keyword}%', *active_dicts))

            results = []
            for row in cursor.fetchall():
                results.append({
                    'word': row[0],
                    'part_of_speech': row[1],
                    'chinese_meaning': row[2],
                    'group_name': row[3],
                    'dict_name': row[4]
                })

            conn.close()
            return results

        except Exception as e:
            print(f"搜索单词失败: {e}")
            return []

    def delete_word_group(self, group_name: str, dict_name: str = None) -> bool:
        """删除单词组"""
        try:
            if not dict_name:
                return False

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                           SELECT wg.id
                           FROM word_groups wg
                                    JOIN dictionaries d ON wg.dict_id = d.id
                           WHERE wg.group_name = ?
                             AND d.dict_name = ?
                           ''', (group_name, dict_name))

            group_result = cursor.fetchone()
            if not group_result:
                conn.close()
                return False

            group_id = group_result[0]

            cursor.execute('DELETE FROM words WHERE group_id = ?', (group_id,))
            cursor.execute('DELETE FROM word_groups WHERE id = ?', (group_id,))

            conn.commit()
            conn.close()
            return True

        except Exception as e:
            print(f"删除单词组失败: {e}")
            return False

    def export_group_to_json(self, group_name: str, file_path: str, dict_name: str = None) -> bool:
        """导出单词组到JSON文件"""
        try:
            if not dict_name:
                return False

            group_data = self.get_word_group(group_name, dict_name)
            if not group_data:
                return False

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(group_data, f, ensure_ascii=False, indent=2)

            return True

        except Exception as e:
            print(f"导出失败: {e}")
            return False

    def import_group_from_json(self, file_path: str, dict_name: str = None) -> bool:
        """从JSON文件导入单词组"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                group_data = json.load(f)

            words_data = []
            for word in group_data.get('words', []):
                words_data.append({
                    'word': word['word'],
                    'part_of_speech': word['part_of_speech'],
                    'chinese_meaning': word['chinese_meaning']
                })

            return self.add_word_group(
                group_data['group_name'],
                words_data,
                group_data.get('description', ''),
                dict_name
            )

        except Exception as e:
            print(f"导入失败: {e}")
            return False

    def get_statistics(self) -> Dict:
        """获取统计信息"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # 总单词数（只统计激活词典）
            active_dicts = self.get_active_dictionaries()
            if active_dicts:
                placeholders = ','.join(['?'] * len(active_dicts))
                cursor.execute(f'''
                    SELECT COUNT(*) FROM words w
                    JOIN word_groups wg ON w.group_id = wg.id
                    JOIN dictionaries d ON wg.dict_id = d.id
                    WHERE d.dict_name IN ({placeholders})
                ''', active_dicts)
                total_words = cursor.fetchone()[0]

                cursor.execute(f'''
                    SELECT COUNT(*) FROM word_groups wg
                    JOIN dictionaries d ON wg.dict_id = d.id
                    WHERE d.dict_name IN ({placeholders})
                ''', active_dicts)
                total_groups = cursor.fetchone()[0]
            else:
                total_words = 0
                total_groups = 0

            # 词典统计
            cursor.execute('''
                           SELECT dict_name,
                                  is_active,
                                  (SELECT COUNT(*) FROM word_groups WHERE dict_id = dictionaries.id) as group_count,
                                  (SELECT COUNT(*)
                                   FROM words w
                                            JOIN word_groups wg ON w.group_id = wg.id
                                   WHERE wg.dict_id = dictionaries.id)                               as word_count
                           FROM dictionaries
                           ORDER BY sort_order
                           ''')
            dict_stats = [
                {
                    'dict_name': row[0],
                    'is_active': bool(row[1]),
                    'group_count': row[2],
                    'word_count': row[3]
                }
                for row in cursor.fetchall()
            ]

            conn.close()

            return {
                'total_words': total_words,
                'total_groups': total_groups,
                'dictionary_statistics': dict_stats,
                'active_dictionaries': active_dicts
            }

        except Exception as e:
            print(f"获取统计信息失败: {e}")
            return {}

    # ========== Excel 导入导出功能 ==========

    def export_to_excel(self, file_path: str) -> bool:
        """导出所有词典到Excel文件"""
        try:
            # 获取所有词典
            dictionaries = self.get_dictionaries_ordered()

            if not dictionaries:
                return False

            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "词汇管理数据库"

            # 添加标题行
            ws.append(["词汇管理数据库", "", ""])  # 第一行第一列
            ws.append(["词典名称", "词组名称", ""])  # 第二行
            ws.append(["英文单词", "词性", "汉译"])  # 第三行

            # 遍历每个词典
            for dict_info in dictionaries:
                dict_name = dict_info['dict_name']

                # 获取该词典的所有单词组
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()

                cursor.execute('''
                    SELECT wg.group_name
                    FROM word_groups wg
                    JOIN dictionaries d ON wg.dict_id = d.id
                    WHERE d.dict_name = ?
                    ORDER BY wg.group_name
                ''', (dict_name,))

                groups = cursor.fetchall()

                # 遍历每个单词组
                for group_row in groups:
                    group_name = group_row[0]

                    # 添加词典和词组信息行（空，词典名称，词组名称）
                    ws.append(["", dict_name, group_name])

                    # 获取该组的单词
                    cursor.execute('''
                        SELECT w.word, w.part_of_speech, w.chinese_meaning
                        FROM words w
                        JOIN word_groups wg ON w.group_id = wg.id
                        JOIN dictionaries d ON wg.dict_id = d.id
                        WHERE d.dict_name = ? AND wg.group_name = ?
                        ORDER BY w.word
                    ''', (dict_name, group_name))

                    words = cursor.fetchall()

                    # 添加单词行
                    for word_row in words:
                        word, pos, meaning = word_row
                        ws.append([word, pos, meaning])

                    # 添加空行分隔单词组
                    ws.append(["", "", ""])

                conn.close()

            # 保存工作簿
            wb.save(file_path)
            return True

        except Exception as e:
            print(f"导出到Excel失败: {e}")
            return False

    
    
    def import_from_excel(self, file_path: str, dict_name: str = None) -> bool:
        """从Excel文件导入词汇数据库"""
        try:
            if dict_name and not self.ensure_dictionary(dict_name):
                return False

            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            current_dict_name = None
            current_group_name = None
            words_data = []
            imported_groups = 0

            def save_current_group():
                nonlocal imported_groups, words_data
                if not (words_data and current_group_name and current_dict_name):
                    return
                target_dict = dict_name if dict_name else current_dict_name
                if not self.ensure_dictionary(target_dict):
                    return
                if self.add_word_group(current_group_name, words_data, "", target_dict):
                    imported_groups += 1
                words_data = []

            for row in ws.iter_rows(min_row=1, values_only=True):
                if all(cell is None for cell in row):
                    continue

                if (row[0] is None or str(row[0]).strip() == "") and row[1]:
                    save_current_group()
                    current_dict_name = str(row[1]).strip()
                    current_group_name = str(row[2]).strip() if row[2] else "未命名组"
                    if not dict_name:
                        self.ensure_dictionary(current_dict_name)

                elif row[0] and row[0] != "词汇管理数据库" and row[0] != "英文单词":
                    word = str(row[0]).strip()
                    pos = str(row[1]).strip() if row[1] else ""
                    meaning = str(row[2]).strip() if row[2] else ""

                    if word and meaning:
                        words_data.append({
                            'word': word,
                            'part_of_speech': pos,
                            'chinese_meaning': meaning
                        })

            save_current_group()

            wb.close()
            return imported_groups > 0

        except Exception as e:
            print(f"从Excel导入失败: {e}")
            return False

    def export_selected_to_excel(self, file_path: str, selected_groups: List[tuple]) -> bool:
        """导出选中的单词组到Excel"""
        try:
            if not selected_groups:
                return False

            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "词汇导出"

            # 添加标题行
            ws.append(["词汇管理数据库", "", ""])
            ws.append(["词典名称", "词组名称", ""])
            ws.append(["英文单词", "词性", "汉译"])

            # 按词典分组
            grouped_by_dict = {}
            for group_name, dict_name in selected_groups:
                if dict_name not in grouped_by_dict:
                    grouped_by_dict[dict_name] = []
                grouped_by_dict[dict_name].append(group_name)

            # 遍历每个词典
            for dict_name, groups in grouped_by_dict.items():
                # 遍历每个单词组
                for group_name in groups:
                    # 获取组数据
                    group_data = self.get_word_group(group_name, dict_name)
                    if not group_data:
                        continue

                    # 添加词典和词组信息行
                    ws.append(["", dict_name, group_name])

                    # 添加单词行
                    for word in group_data['words']:
                        ws.append([word['word'], word['part_of_speech'], word['chinese_meaning']])

                    # 添加空行分隔
                    ws.append(["", "", ""])

            # 保存工作簿
            wb.save(file_path)
            return True

        except Exception as e:
            print(f"导出选中组到Excel失败: {e}")
            return False